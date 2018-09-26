import yaml
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from string import ascii_uppercase
from osgeo import ogr, osr
from re import findall


__author__ = 'Jacob Wilson (JNCC)'


def break_down_fields(field_list):
    """
    Breaks a nested list down into its constituents all in a flat list
    :param field_list: Top level of nested list
    :return: Flat list
    """
    if not isinstance(field_list, (list, tuple, set)):
        return [field_list]
    broken_down = []
    for item in field_list:
        if isinstance(item, (list, tuple, set)):
            sub_fields = break_down_fields(item)
            for sub_field in sub_fields:
                broken_down.append(sub_field)
        else:
            broken_down.append(item)
    return broken_down


def check_all_arguments_are_floats(f):
    """
    Decorates a function by checking that all its arguments are floats or lists of floats
    :param f: Function
    :return: None if not all arguments are floats
    """
    def checker(x):
        x = break_down_fields(x)
        try:
            for index in range(len(x)):
                x[index] = float(x[index])
        except (ValueError, TypeError):
            return None
        else:
            return f(x)
    return checker


def turn_data_to_args(f):
    """
    Decorates a function by turning a list given to it into separate arguments
    :param f: Function
    """
    def wrapper(data_in):
        return f(*data_in)
    return wrapper


def import_config(path):
    with open(path, 'r', encoding='utf-8') as stream:
        return yaml.load(stream)


def get_data_from_field(field_names):
    """

    :param field_names:
    :return:
    """
    content = []
    for field_name in break_down_fields(field_names):
        content.append(row[FIELD_TO_INDEX[field_name]])
    return content


def get_indices_for_fields(sheet, fields):
    """
    Relates the field names to their indices in the spreadsheet
    :param sheet: Spreadsheet
    :param fields: Field names
    :return: Dictionary relating indices and field names
    """
    relations = {}
    for field_name in break_down_fields(fields):
        if field_name is None:
            pass
        try:
            relations[field_name] = [sheet_field.strip('\n').strip() for sheet_field in list(sheet.values)[config['Analysis_Workbook']['Field_Row'] - 1] if sheet_field is not None].index(field_name)
        except ValueError:
            pass
    return relations


def mandatory_check(value):
    """
    Checks if a mandatory value exists
    """
    assert value is not None, 'Mandatory value not found: {}, {}'.format(value, field)
    return value


def iter_rows_until_blank(sheet, min_row=0, max_row=float('inf')):
    """
    Iterates through a spreadsheet's rows until it hits a row with a blank A column value
    :param sheet: Spreadsheet
    :param min_row: Row to start iterating from (inclusive)
    :param max_row: Row to end iterating at (inclusive)
    :yield: Each row
    """
    count = min_row - 1
    values = list(sheet.values)
    while count <= max_row:
        try:
            current_row = values[count]
        except IndexError:
            raise StopIteration
        if current_row[0] is None:
            raise StopIteration
        yield current_row
        count += 1


@check_all_arguments_are_floats
@turn_data_to_args
def reproject_eastings_to_longitude(eastings, northings, srcepsgcode):
    """
    Reprojects an OSGB E/N position to a lat/long using WGS84
    :param eastings: Easting of OSGB point
    :param northings: Northing of OSGB point
    :return: Longitude of transformed point
    """

    source = osr.SpatialReference()
    source.ImportFromEPSG(srcepsgcode)

    target = osr.SpatialReference()
    target.ImportFromEPSG(4326)

    transform = osr.CoordinateTransformation(source, target)

    point = ogr.CreateGeometryFromWkt("POINT ({} {})".format(eastings, northings))
    point.Transform(transform)

    longitude = str(point.ExportToWkt()).split()[2].strip(')')

    return [abs(float(longitude)), float(longitude)]


@check_all_arguments_are_floats
@turn_data_to_args
def reproject_northings_to_latitude(eastings, northings):
    """
    Reprojects an OSGB E/N position to a lat/long using WGS84
    :param eastings: Easting of OSGB point
    :param northings: Northing of OSGB point
    :return: Latitude of transformed point
    """

    source = osr.SpatialReference()
    source.ImportFromEPSG(32631)

    target = osr.SpatialReference()
    target.ImportFromEPSG(4326)

    transform = osr.CoordinateTransformation(source, target)

    point = ogr.CreateGeometryFromWkt("POINT ({} {})".format(eastings, northings))
    point.Transform(transform)

    latitude = str(point.ExportToWkt()).split()[1].strip('(')

    return [abs(float(latitude)), float(latitude)]


@check_all_arguments_are_floats
def latitude_direction(latitude):
    if latitude[0] >= 0:
        return 'N'
    if latitude[0] < 0:
        return 'S'


@check_all_arguments_are_floats
def longitude_direction(longitude):
    if longitude[0] >= 0:
        return 'E'
    if longitude[0] < 0:
        return 'W'

@check_all_arguments_are_floats
def get_absolute_value(number):
    return abs(number[0])

def split_EUNIS(code):
    """
    Splits up a EUNIS code into each of it hierarchical layers
    :param code: EUNIS code to split
    :return: Split up code
    """
    code = code[0]
    try:
        assert isinstance(code, str), 'EUNIS code not understood'
    except AssertionError:
        return None
    split = ''
    cache = ''
    for character in code:
        if character == '.':
            cache += character
        else:
            cache += character
            if split != '':
                split += '|'
            split += cache
    return split


def split_MNCR(code):
    """
    Splits up a MNCR code into each of it hierarchical layers
    :param code: MNCR code to split
    :return: Split up code
    """
    code = code[0]
    try:
        assert isinstance(code, str), 'MNCR code not understood'
    except AssertionError:
        return None
    split = ''
    cache = ''
    for group in code.split('.'):
        if cache != '':
            cache += '.'
        cache += group
        if split != '':
            split += '|'
        split += cache
    return split


@turn_data_to_args
def find_regex_matches(field_for_string, pattern):
    """
    Returns all regex matches from pattern in the string found in the given field
    :param field_for_string: Field to take string from
    :param pattern: Regex pattern to match against
    :return: Matches found
    """
    regex_string = get_data_from_field(field_for_string)[0]
    return findall(pattern, regex_string)


def get_indices_for_noted_species():
    """
    Creates a dictionary relating indices in the species matrix to the noted species
    :return: Dictionary (index: species)
    """
    list_of_noted_species = [cell.value.lower() for cell in species_of_note_workbook['A']][1:]
    relations = {}
    values = [i[config['Species_Matrix']['Species_Name_Column_Number'] - 1] for i in list(species_matrix.values)][1:]
    for species_index, species in enumerate(values):
        if species is None:
            continue
        if species.lower() in list_of_noted_species:
            relations[species_index + 1] = species
    return relations


def get_species_of_note(data):
    """
    Returns any species in the species of note list that are found in a given column
    """
    noted_species_present = ''
    for column_index in range(len(species_matrix_values[0]) - config['Species_Matrix']['Species_Name_Column_Number']):
        if get_data_from_field('Still Sample Ref')[0] == species_matrix_values[0][column_index]:
            for species_of_note_index in INDEX_TO_SPECIES_OF_NOTE:
                if species_matrix_values[species_of_note_index][column_index] != config['Species_Matrix']['Blank_Entry_Looks_Like']:
                    noted_species_present += INDEX_TO_SPECIES_OF_NOTE[species_of_note_index] + '|'
    return noted_species_present[:-1] if noted_species_present != '' else None


def common_name_list(species_of_note_field_name):
    """
    Returns the common names of the species of note
    :param species_of_note_field_name: Field name for the specie of note
    :return: List of common names separated by pipes '|'
    """
    taxa_list = current_row[species_of_note_field_name]
    if taxa_list is None:
        return None
    if '|' in taxa_list:
        taxa_list = taxa_list.split('|')
    else:
        taxa_list = [taxa_list]

    common_names = ''
    for taxa in taxa_list:
        for species_row in species_of_note_workbook.values:
            if species_row[0] == taxa:
                if species_row[1] is None:
                    continue
                common_names += '|' + species_row[1]

    return common_names[1:]


# Relates function names from config to functions in code, each must take the analysis, row and the data from config as parameters
NAME_TO_FUNCTION = {
    'mandatory check': mandatory_check,
    'get absolute value': get_absolute_value,
    'get data from field': get_data_from_field,
    'reproject eastings': reproject_eastings_to_longitude,
    'reproject northings': reproject_northings_to_latitude,
    'get latitude direction': latitude_direction,
    'get longitude direction': longitude_direction,
    'handle eunis code': split_EUNIS,
    'handle mncr code': split_MNCR,
    'get species of note': get_species_of_note,
    'find regex matches': find_regex_matches,
    'absolute value': check_all_arguments_are_floats(turn_data_to_args(abs)),
    'get common name of species': common_name_list
}


def convert_from_number_to_excel_letter(number):
    """
    Takes a number and turns it into an excel horizontal index
    E.g. 3 -> C ;  27 -> AA
    :param number: Number to be converted
    :return: Index made out of letters
    """
    letter = ''
    highest_power_of_26 = 0
    while (number // (26 ** highest_power_of_26)) > 26:
        highest_power_of_26 += 1
    for power in range(highest_power_of_26, -1, -1):
        floor = number // 26 ** power
        letter += ascii_uppercase[floor - 1]
        number -= floor * 26

    return letter


if __name__ == '__main__':
    # Load config and spreadsheets
    ## TODO  - Handle yaml file drops or else prompt user
    config = import_config(r'Z:\Marine\Evidence\BenthicDataMgmt\Image catalogue\batch_tagging_tool\1714s_pilot_2\1714s.yaml')
    analysis = load_workbook(config['Analysis_Workbook']['Path'])[config['Analysis_Workbook']['Sheet_Name']]
    species_matrix = load_workbook(config['Species_Matrix']['Path'])[config['Species_Matrix']['Sheet_Name']]
    species_of_note_workbook = load_workbook(config['Species_of_Note_Workbook']).active
    species_matrix_values = list(species_matrix.values)

    # Create a new workbook to hold the outputs
    new_metadata_workbook = Workbook()
    new_sheet = new_metadata_workbook.active
    # Put fields into new workbook
    for index, field in enumerate(config['Fields']):
        new_sheet[get_column_letter(index + 1) + '1'] = field

    # Produce dictionary relating indices in the species matrix to any noted species in it
    INDEX_TO_SPECIES_OF_NOTE = get_indices_for_noted_species()
    # Produce dictionary relating each field name to its horizontal index in the spreadsheet
    FIELD_TO_INDEX = get_indices_for_fields(analysis, [input_field[1] for input_field in config['Fields'].values()])
    # Iterate over each image
    for row_index, row in enumerate(iter_rows_until_blank(analysis, config['Analysis_Workbook']['Data_Starts_at_Row_Number'])):
        current_row = {}
        # Create a new value for the fields specified in config
        for column_index, field in enumerate(config['Fields']):
            # Take data from config
            data = config['Fields'][field][1]
            for function_to_be_run in config['Fields'][field][0]:
                # Run function specified in config
                data = NAME_TO_FUNCTION[function_to_be_run.lower()](data)
            if isinstance(data, list):
                data = data[0]
            try:
                data = float(data)
            except (TypeError, ValueError):
                pass
            current_row[field] = data
            # Output to final sheet
            new_sheet[convert_from_number_to_excel_letter(column_index + 1) + str(row_index + 2)] = data
    # Save output file
    new_metadata_workbook.save(config['File_Name'])
