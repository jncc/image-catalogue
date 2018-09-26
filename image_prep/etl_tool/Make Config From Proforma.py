from openpyxl import load_workbook
from collections import OrderedDict
import oyaml
import yaml.representer
from pathlib import Path, PurePath


def iter_rows_until_blank(sheet, min_row=0, max_row=float('inf')):
    """
    Iterates through a spreadsheet's rows until it hits a row with a blank A column value
    :param sheet: Spreadsheet
    :param min_row: Row to start iterating from (inclusive)
    :param max_row: Row to end iterating at (inclusive)
    :yield: Each row
    """
    count = min_row - 1
    values = list(sheet.values)
    while count <= max_row:
        try:
            current_row = values[count]
        except IndexError:
            raise StopIteration
        if current_row[0] is None:
            raise StopIteration
        yield current_row
        count += 1

config = OrderedDict({'File_Name': None,
          'Species_of_Note_Workbook': None,
          'Analysis_Workbook':
              {'Path': None,
               'Sheet_Name': None,
               'Field_Row': None,
               'Data_Starts_at_Row_Number': None},
          'Species_Matrix':
              {'Path': None,
               'Sheet_Name': None,
               'Species_Name_Column_Number': None,
               'Blank_Entry_Looks_Like': None},
          'Fields': {}
          })

def literal_presenter(dumper, data):
    if isinstance(data, str) and "\n" in data:
        return dumper.represent_scalar('tag:yaml.org,2002:str', data, style='|')
    return dumper.represent_scalar('tag:yaml.org,2002:str', data, style='"')


oyaml.add_representer(str, literal_presenter)

if __name__ == '__main__':
    OUTPUT_FILE_NAME = Path(input('Type output directory for config: ')) / 'etlconfig.yaml'

    PROFORMA_PATH = str(Path(__file__).parents[2] / 'reference_templates' / 'image_proforma_template.xlsx')

    proforma = load_workbook(PROFORMA_PATH, data_only=True)['Fields_Descriptions']

    # Iterates through proforma adding the field names to a new config
    for row in iter_rows_until_blank(proforma, 3):
        config['Fields'][row[0]] = [[], ""]

    with open(OUTPUT_FILE_NAME, 'w') as outfile:
        yaml.dump(config, outfile, default_flow_style=False)
